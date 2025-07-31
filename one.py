from openpyxl import load_workbook
from datetime import datetime
workbook= load_workbook(r"C:\Users\Srikrishna Vadlamani\OneDrive\Desktop\Book1.xlsx")
sheet = workbook["Sheet1"]
updates = {
    "x" : 17,
    "y" : 81,
    "z" : 91
}
updated=[]
for row in sheet.iter_rows():
    for cell in row:
        key = str(cell.value).strip() if cell.value is not None else ""
        if key in updates:
            next_col = cell.column + 1
            sheet.cell(row=cell.row,column=next_col,value=updates[key])
            print(f"Updated cell at {cell.row}, {next_col}")
            updated.append(key)
now = datetime.now()
formatted_time = now.strftime("%b%d_%H-%M-%S")
new_file = f"{formatted_time}.xlsx"
path = fr"C:\Users\Srikrishna Vadlamani\OneDrive\Desktop\{new_file}"
workbook.save(path)
if updated:
    print(updated)
else:
    print("No matching variables found")