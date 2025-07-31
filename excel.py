from flask import send_file
from openpyxl import load_workbook
from datetime import datetime
import io
path = r"C:\Users\Srikrishna Vadlamani\OneDrive\Desktop\Book1.xlsx"
headers = ["Emp_id","Emp_name","Email","Password"]
def generate_row(data):
    wb= load_workbook(path)
    sheet = wb.active
    header_row = None
    for row in sheet.iter_rows(min_row=1,max_row=30):
        values = [str(cell.value).strip() if cell.value else "" for cell in row]
        if all(h in values for h in headers):
            header_row = row[0].row
            break
    if header_row is None:
        return "Header row not found!", 400
    col_map = {}
    for col_idx,cell in enumerate(sheet[header_row],start=1):
        if cell.value in headers:
            col_map[cell.value] = col_idx
    current_row = header_row + 1
    while any(sheet.cell(row=current_row,column=col).value for col in col_map.values()):
        current_row += 1
    sheet.cell(row=current_row,column=col_map["Emp_name"],value=data['emp_name'])
    sheet.cell(row=current_row,column=col_map["Emp_id"],value=data['emp_id'])
    sheet.cell(row=current_row,column=col_map["Email"],value=data['email'])
    sheet.cell(row=current_row,column=col_map["Password"],value=data['password'])
    return generate_response(wb)
def generate_column(data):
    wb= load_workbook(path)
    sheet = wb.active
    header_col = None
    for col in sheet.iter_cols(min_col=1,max_col=30):
        values = [str(cell.value).strip() if cell.value else "" for cell in col]
        if all(h in values for h in headers):
            header_col = col[0].column
            break
    if header_col is None:
        return "Header col not found!",401
    row_map = {}
    for row in sheet.iter_rows(min_col=header_col,max_col=header_col):
        for cell in row:
            if cell.value in headers:
                row_map[cell.value]=cell.row
    current_col = header_col + 1
    while any(sheet.cell(row=row,column=current_col).value for row in row_map.values()):
        current_col += 1
    sheet.cell(row=row_map["Emp_id"],column=current_col,value=data['emp_id'])
    sheet.cell(row=row_map["Emp_name"],column=current_col,value=data['emp_name'])
    sheet.cell(row=row_map["Email"],column=current_col,value=data['email'])
    sheet.cell(row=row_map["Password"],column=current_col,value=data['password'])
    return generate_response(wb)
def generate_response(workbook):
    file_stream = io.BytesIO()
    time = datetime.now().strftime("%b%d_%H-%M-%S")
    file_name = f"{time}.xlsx"
    workbook.save(file_stream)
    file_stream.seek(0)
    return send_file(file_stream,as_attachment=True,download_name=file_name,
    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')